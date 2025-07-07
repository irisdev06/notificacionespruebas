import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import numpy as np
from itertools import cycle, islice
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill

def cargar_archivo(archivo, tipo):
    try:
        if tipo == "xlsx":
            # Cargar ambas hojas (DTO y PCL)
            df_dto = pd.read_excel(archivo, sheet_name='DTO')
            df_pcl = pd.read_excel(archivo, sheet_name='PCL')

            # Unir ambas hojas en un solo DataFrame
            df_base = pd.concat([df_dto, df_pcl], ignore_index=True)

        elif tipo == "csv":
            df_base = pd.read_csv(archivo, on_bad_lines='skip', delimiter=",")  # 'skip' ignora las líneas mal formadas

        # Limpiar posibles filas con datos inconsistentes
        df_base.dropna(how='all', inplace=True)  # Eliminar filas vacías
        df_base = df_base.reset_index(drop=True)  # Resetear el índice

        return df_base
    except Exception as e:
        st.error(f"Error al procesar el archivo {tipo}: {e}")
        return None

def grafica_barras(df_base, workbook):
    # Verificar columnas necesarias
    if 'ESTADO_INFORME' not in df_base.columns or 'NOTIFICADOR' not in df_base.columns:
        st.error("El archivo no contiene las columnas necesarias: 'ESTADO_INFORME' y 'NOTIFICADOR'.")
        return workbook

    # Agrupar datos
    conteo = df_base.groupby(['ESTADO_INFORME', 'NOTIFICADOR']).size().unstack(fill_value=0)

    estados = conteo.index
    notificadores = conteo.columns
    x = np.arange(len(estados))

    colores = ['#809bce', '#95b8d1', "#79cbd1", '#B8E6A7', '#4C9A2A']
    colores_usar = list(islice(cycle(colores), len(notificadores)))

    total_width = 0.8
    bar_width = total_width / len(notificadores)

    fig, ax = plt.subplots(figsize=(max(15, len(estados) * 0.4), 6))

    for i, notificador in enumerate(notificadores):
        bars = ax.bar(x + i * bar_width, conteo[notificador], width=bar_width, label=notificador, color=colores_usar[i])
        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, yval, int(yval), ha='center', va='bottom', fontsize=8)

    ax.set_xticks(x + total_width / 2 - bar_width / 2)
    ax.set_xticklabels(estados, rotation=90, ha='center', fontsize=7)
    ax.set_xlabel('Estado de Informe')
    ax.set_ylabel('Cantidad')
    ax.set_title('Distribución de Notificadores por Estado de Informe')
    ax.legend(title='Notificadores', bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.tight_layout()

    # Guardar la figura como imagen con fondo transparente
    imgdata = BytesIO()
    plt.savefig(imgdata, format='png', dpi=200, transparent=True)  # <- AQUÍ el cambio
    plt.close()
    imgdata.seek(0)

    # Crear hoja nueva
    if 'Distribución de Notificadores' in [s.title for s in workbook.worksheets]:
        sheet = workbook['Distribución de Notificadores']
    else:
        sheet = workbook.create_sheet('Distribución de Notificadores')

    # Insertar la imagen usando openpyxl
    imagen = ExcelImage(imgdata)
    imagen.anchor = 'A1'
    sheet.add_image(imagen)

    return workbook


# -------------------------- FUNCIONES DE PROCESAMIENTO Y GENERACIÓN DE TABLAS ---------------------------
def generar_tablas_estado_informe(archivo, tipo):
    df_base = cargar_archivo(archivo, tipo)

    if df_base is not None:
        if "ESTADO_INFORME" in df_base.columns and "NOTIFICADOR" in df_base.columns:
            conteo = df_base.groupby(['ESTADO_INFORME', 'NOTIFICADOR']).size().unstack(fill_value=0)
            conteo['TOTAL GENERAL'] = conteo.sum(axis=1)
            total_general_sum = conteo['TOTAL GENERAL'].sum()
            conteo['TOTAL %'] = (conteo['TOTAL GENERAL'] / total_general_sum) * 100
            conteo['TOTAL %'] = conteo['TOTAL %'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else '')

            libro = Workbook()
            hoja_procesada = libro.active
            hoja_procesada.title = "Tabla Procesada"

            hoja_procesada.cell(row=1, column=1, value="ESTADO INFORME")
            for col_idx, notificador in enumerate(conteo.columns[:-2], start=2):
                hoja_procesada.cell(row=1, column=col_idx, value=notificador)
            hoja_procesada.cell(row=1, column=len(conteo.columns)-1, value="TOTAL GENERAL")
            hoja_procesada.cell(row=1, column=len(conteo.columns), value="TOTAL %")

            row_start = 2
            for estado, valores in conteo.iterrows():
                hoja_procesada.cell(row=row_start, column=1, value=estado)
                for col_idx, notificador in enumerate(conteo.columns[:-2], start=2):
                    hoja_procesada.cell(row=row_start, column=col_idx, value=valores.get(notificador, 0))
                hoja_procesada.cell(row=row_start, column=len(valores)-2, value=valores.get('TOTAL GENERAL', 0))
                hoja_procesada.cell(row=row_start, column=len(valores)-1, value=valores.get('TOTAL %', ''))
                row_start += 1

            hoja_procesada.cell(row=row_start, column=1, value="TOTAL GENERAL")
            for col_idx in range(2, len(conteo.columns)-1):
                hoja_procesada.cell(row=row_start, column=col_idx, value=conteo.iloc[:, col_idx - 1].sum())
            hoja_procesada.cell(row=row_start, column=len(conteo.columns)-1, value='')

            hoja_base = libro.create_sheet("BASE")
            for r_idx, row in enumerate(dataframe_to_rows(df_base, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    hoja_base.cell(row=r_idx, column=c_idx, value=value)

            # Estilos
            borde = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )
            fondo_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            fondo_gris_total = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

            for cell in hoja_procesada["1:1"]:
                cell.border = borde
                cell.fill = fondo_gris
            for cell in hoja_procesada[row_start]:
                cell.fill = fondo_gris_total
                cell.border = borde
            for row in hoja_procesada.iter_rows(min_row=2, max_row=row_start, min_col=1, max_col=len(conteo.columns)):
                for cell in row:
                    cell.border = borde

            # Ajuste ancho
            for col in hoja_procesada.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                hoja_procesada.column_dimensions[column].width = max_length + 2

            # AGREGAR LA GRÁFICA
            libro = grafica_barras(df_base, libro)

            # GUARDAR EL ARCHIVO
            output = BytesIO()
            libro.save(output)
            output.seek(0)
            return output

        else:
            st.error("El archivo no contiene las columnas necesarias: 'ESTADO_INFORME' y 'NOTIFICADOR'.")
            return None


# ------------------------ FUNCIONES DE SUBIDA Y DESCARGA -------------------------------

# Función para descargar el archivo generado
def descargar_excel(output, nombre="archivo_procesado.xlsx"):
    st.download_button(
        label="📥 Descargar archivo",  
        data=output,
        file_name=nombre,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Función para subir el archivo
def subir_archivo2():
    archivo = st.file_uploader("Sube un archivo (.xlsx o .csv)", type=["xlsx", "csv"], key="file_uploader")
    
    if archivo is not None:
        try:
            nombre_archivo = archivo.name.lower()

            if nombre_archivo.endswith(".xlsx"):
                st.success("¡Archivo Excel válido!")
                return archivo, "xlsx"
            elif nombre_archivo.endswith(".csv"):
                st.success("¡Archivo CSV válido!")
                return archivo, "csv"
            else:
                st.warning("El archivo debe ser de tipo .xlsx o .csv")
                return None, None

        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")
            return None, None

    return None, None


# ---------------------------- FLUJO  --------------------------

# Función para procesar el archivo y generar la tabla
def procesar_archivos2():
    archivo, tipo = subir_archivo2()

    if archivo and tipo in ["xlsx", "csv"]:
        # Cargar y procesar el archivo
        df_base = cargar_archivo(archivo, tipo)
        
        if df_base is not None:
            # Generar las tablas
            output = generar_tablas_estado_informe(archivo, tipo)

            # Crear el libro de trabajo (workbook) vacío
            libro = Workbook()

            # Llamar a la función de la gráfica
            libro = grafica_barras(df_base, libro)

            if output:
                # Descarga el archivo generado
                descargar_excel(output, nombre="informe_estado_informe.xlsx")
                st.success("✅ Archivo generado con éxito con el gráfico.")
        else:
            st.error("El archivo no contiene datos válidos para procesar.")
    else:
        st.error("No se ha cargado un archivo válido.")
