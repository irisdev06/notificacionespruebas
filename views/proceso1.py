import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill
import calendar
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import csv

# Colores para las gráficas
colores = ['#FFB897', '#B8E6A7', '#809bce', "#64a09d", '#CBE6FF', '#E6E6FA']

# ------------------------------------------------------------------------------- GRÁFICOS DE BARRAS -------------------------------------------------------------
def graficas_barras(df, colores, nombre_hoja):
    conteo = df.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    conteo.index = conteo.index.map(lambda m: calendar.month_name[m].capitalize())
    num_meses = len(conteo)
    num_notificadores = len(conteo.columns)
    fig_width = max(12, num_meses * 1.2)
    fig_height = max(8, num_notificadores * 1.0)
    ax = conteo.plot(kind='bar', figsize=(fig_width, fig_height), color=colores)
    ax.set_xlabel('Mes')
    ax.set_ylabel('Número de Datos')
    ax.legend(title='NOTIFICADOR', bbox_to_anchor=(1.2, 1), loc='upper left', fontsize=10)

    for p in ax.patches:
        ax.annotate(f'{p.get_height()}', 
                    (p.get_x() + p.get_width() / 2., p.get_height()), 
                    xytext=(0, 5),
                    textcoords='offset points',
                    ha='center', va='bottom', fontsize=10, color='black')

    grafico_path = f"{nombre_hoja}_grafico_barras.png"
    ax.get_figure().savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path


def graficas_barras_comparativa(df, nombre_hoja):
    # Filtrar solo los datos de BELISARIO397 y GESTAR INNOVACION
    df_comparativa = df[df['NOTIFICADOR'].isin(['BELISARIO397', 'GESTAR INNOVACION'])]
    
    conteo = df_comparativa.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    conteo.index = conteo.index.map(lambda m: calendar.month_name[m].capitalize())
    
    # Crear la gráfica de barras
    fig, ax = plt.subplots(figsize=(12, 8))
    conteo.plot(kind='bar', ax=ax, color=colores)
    ax.set_xlabel('Mes')
    ax.set_ylabel('Número de Datos')
    ax.legend(title='Notificadores', bbox_to_anchor=(1.2, 1), loc='upper left', fontsize=10)

    for p in ax.patches:
        ax.annotate(f'{p.get_height()}', 
                    (p.get_x() + p.get_width() / 2., p.get_height()), 
                    xytext=(0, 5),
                    textcoords='offset points',
                    ha='center', va='bottom', fontsize=10, color='black')

    grafico_path = f"{nombre_hoja}_grafico_barras_comparativa.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path

def graficas_barras_belisario_utmdl(df, nombre_hoja, mes):
    # Filtrar solo los datos de BELISARIO y UTMDL para el mes seleccionado
    df_filtrado = df[(df['NOTIFICADOR'].isin(['BELISARIO', 'UTMDL'])) & (df['MES'] == mes)]
    
    conteo = df_filtrado.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    conteo.index = conteo.index.map(lambda m: calendar.month_name[m].capitalize())
    
    # Crear la gráfica de barras
    fig, ax = plt.subplots(figsize=(12, 8))
    conteo.plot(kind='bar', ax=ax, color=colores)
    ax.set_xlabel('Mes')
    ax.set_ylabel('Número de Datos')
    ax.legend(title='Notificadores', bbox_to_anchor=(1.2, 1), loc='upper left', fontsize=10)

    for p in ax.patches:
        ax.annotate(f'{p.get_height()}', 
                    (p.get_x() + p.get_width() / 2., p.get_height()), 
                    xytext=(0, 5),
                    textcoords='offset points',
                    ha='center', va='bottom', fontsize=10, color='black')

    grafico_path = f"{nombre_hoja}_grafico_barras_belisario_utmdl_{mes}.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path
# ------------------------------------------------------------------------------- GRÁFICOS DE PASTEL -------------------------------------------------------------
def graficas_pastel(df, nombre_hoja):
    conteo = df.groupby('MES').size()
    conteo.index = conteo.index.map(lambda m: calendar.month_name[m].capitalize())  
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(conteo, labels=conteo.index, autopct='%1.1f%%', startangle=90, colors=colores)
    ax.legend(title='Meses', loc='center left', bbox_to_anchor=(1.05, 0.5), fontsize=10)

    grafico_path = f"{nombre_hoja}_grafico_pastel.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path

def graficas_pastel_belisario_utmdl(df, nombre_hoja):
    # Agrupar por NOTIFICADOR (BELISARIO y UTMDL)
    conteo = df[df['NOTIFICADOR'].isin(['BELISARIO', 'UTMDL'])].groupby('NOTIFICADOR').size()
    
    # Crear una gráfica de pastel para BELISARIO y UTMDL
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(conteo, labels=conteo.index, autopct='%1.1f%%', startangle=90, colors=colores)
    ax.legend(title='Notificadores', loc='center left', bbox_to_anchor=(1.05, 0.5), fontsize=10)

    grafico_path = f"{nombre_hoja}_grafico_pastel_belisario_utmdl.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path

# ------------------------------------------------------------------------------- HOJAS -------------------------------------------------------------
def crear_hoja_mes_seleccionado(libro, nombre_hoja, df, mes):
    # Asegurarse de que la columna 'MES' esté presente en el DataFrame antes de filtrar
    df['MES'] = df['FECHA_VISADO'].dt.month
    
    # Filtrar los datos por el mes seleccionado
    df_mes = df[df['MES'] == mes]
    
    # Crear la hoja en el libro
    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]
    hoja = libro.create_sheet(nombre_hoja)

    # Escribir los datos filtrados
    for i, row in enumerate(dataframe_to_rows(df_mes, index=False, header=True), start=1):
        for j, value in enumerate(row, start=1):
            hoja.cell(row=i, column=j, value=value)

    # Generar gráficos de barras y pastel por mes
    grafico_barras_belisario_utmdl_path = graficas_barras_belisario_utmdl(df, nombre_hoja, mes)  # Ahora pasa el mes
    img_barras_belisario_utmdl = Image(grafico_barras_belisario_utmdl_path)
    hoja.add_image(img_barras_belisario_utmdl, 'E5')

    # Generar gráfica de pastel para BELISARIO y UTMDL
    grafico_belisario_utmdl_path = graficas_pastel_belisario_utmdl(df, nombre_hoja)
    img_belisario_utmdl = Image(grafico_belisario_utmdl_path)
    hoja.add_image(img_belisario_utmdl, 'E35')  # Colocar la imagen más abajo en la hoja

def crear_comparativa_ano(libro, df):
    # Crear la hoja "COMPARATIVA AÑO"
    if "COMPARATIVA AÑO" in libro.sheetnames:
        del libro["COMPARATIVA AÑO"]
    hoja = libro.create_sheet("COMPARATIVA AÑO")

    # Filtrar solo los datos de BELISARIO397 y GESTAR INNOVACION
    df_comparativa = df[df['NOTIFICADOR'].isin(['BELISARIO397', 'GESTAR INNOVACION'])]

    # Crear la tabla comparativa por mes
    tabla_comparativa = df_comparativa.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    tabla_comparativa.index = tabla_comparativa.index.map(lambda m: calendar.month_name[m].capitalize())

    # Escribir la tabla comparativa en la hoja
    for i, row in enumerate(dataframe_to_rows(tabla_comparativa, index=True, header=True), start=1):
        for j, value in enumerate(row, start=1):
            hoja.cell(row=i, column=j, value=value)

    # Generar el gráfico de barras comparativo
    grafico_barras_comparativa_path = graficas_barras_comparativa(df_comparativa, "COMPARATIVA AÑO")
    img_comparativa_barras = Image(grafico_barras_comparativa_path)
    hoja.add_image(img_comparativa_barras, 'E5')

    # Generar gráfico de pastel comparativo
    grafico_pastel_comparativa_path = graficas_pastel_belisario_utmdl(df_comparativa, "COMPARATIVA AÑO")
    img_comparativa_pastel = Image(grafico_pastel_comparativa_path)
    hoja.add_image(img_comparativa_pastel, 'E20')

# ------------------------------------------------------------------------------- GENERAR TABLAS Y GRÁFICOS PARA DTO Y PCL -------------------------------------------------------------
def generar_tablas_dto_y_pcl(libro, df_dto, df_pcl):
    def crear_hoja(nombre_hoja, df):
        # Asegurarse de que la columna 'MES' esté presente en el DataFrame antes de realizar el agrupamiento
        df['MES'] = df['FECHA_VISADO'].dt.month
        conteo = df.groupby('MES').size().reset_index(name='TOTAL')
        conteo['MES'] = conteo['MES'].apply(lambda m: calendar.month_name[m].capitalize())
        total_general = conteo['TOTAL'].sum()
        conteo['PORCENTAJE'] = (conteo['TOTAL'] / total_general * 100).round(2).astype(str) + '%'

        fila_total = pd.DataFrame({
            'MES': ['Total general'],
            'TOTAL': [total_general],
            'PORCENTAJE': ['100.0%']
        })
        tabla_final = pd.concat([conteo, fila_total], ignore_index=True)

        if nombre_hoja in libro.sheetnames:
            del libro[nombre_hoja]
        hoja = libro.create_sheet(nombre_hoja)

        borde = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        fondo_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        hoja['A1'] = "FECHA VISADO"
        hoja['B1'] = "TOTAL"
        hoja['C1'] = "PORCENTAJE"

        for celda in ['A1', 'B1', 'C1']:
            hoja[celda].border = borde
            hoja[celda].fill = fondo_gris

        for i, fila in enumerate(dataframe_to_rows(tabla_final, index=False, header=False), start=2):
            hoja[f'A{i}'] = fila[0]
            hoja[f'B{i}'] = fila[1]
            hoja[f'C{i}'] = fila[2]

            hoja[f'A{i}'].border = borde
            hoja[f'B{i}'].border = borde
            hoja[f'C{i}'].border = borde

            if fila[0] == "Total general":
                hoja[f'A{i}'].fill = fondo_gris
                hoja[f'B{i}'].fill = fondo_gris
                hoja[f'C{i}'].fill = fondo_gris

        # Generar gráficos
        grafico_barras_path = graficas_barras(df, colores, nombre_hoja)
        img_barras = Image(grafico_barras_path)
        hoja.add_image(img_barras, 'E5')

        grafico_pastel_path = graficas_pastel(df, nombre_hoja)
        img_pastel = Image(grafico_pastel_path)
        hoja.add_image(img_pastel, 'E20')

    # Crear las hojas para DTO y PCL
    crear_hoja("TABLA MES DTO", df_dto)
    crear_hoja("TABLA MES PCL", df_pcl)

# ------------------------------------------------------------------------------- FUNCIONES DE SUBIDA Y DESCARGA -------------------------------------------------------------
def descargar_archivo(output, nombre="archivo_procesado.xlsx"):
    st.download_button(
        label="📥 Descargar archivo",  
        data=output,
        file_name=nombre,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def subir_archivo():
    archivo = st.file_uploader("Sube un archivo (.xlsx o .csv)", type=["xlsx", "csv"])

    if archivo is not None:
        try:
            nombre_archivo = archivo.name.lower()

            if nombre_archivo.endswith(".xlsx"):
                xls = pd.ExcelFile(archivo)
                hojas = xls.sheet_names

                if "DTO" in hojas and "PCL" in hojas:
                    st.success("¡Archivo Excel válido! Se encontraron las hojas DTO y PCL.")
                    return archivo, "xlsx"
                else:
                    if "DTO" not in hojas:
                        st.error("La hoja 'DTO' no se encuentra en el archivo.")
                    if "PCL" not in hojas:
                        st.error("La hoja 'PCL' no se encuentra en el archivo.")
                    return None, None

            elif nombre_archivo.endswith(".csv"):
                df = pd.read_csv(archivo)
                if "DTO" in df.columns and "PCL" in df.columns:
                    st.success("¡Archivo CSV válido! Se encontraron las columnas DTO y PCL.")
                    return archivo, "csv"
                else:
                    st.warning("El archivo CSV no contiene columnas llamadas 'DTO' y 'PCL'.")
                    return None, None

        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")
            return None, None

    return None, None


# ------------------------------------------------------------------------------- FLUJO ---------------------------------------------------------------------------------
def procesar_archivos():
    archivo, tipo = subir_archivo()

    if archivo and tipo == "xlsx":
        # Mostrar el selector de mes
        mes_seleccionado = st.selectbox("Selecciona el mes", list(calendar.month_name[1:]))  # Los meses son de 1 a 12

        # Leer las hojas DTO y PCL
        df_dto = pd.read_excel(archivo, sheet_name='DTO', parse_dates=['FECHA_VISADO'])
        df_pcl = pd.read_excel(archivo, sheet_name='PCL', parse_dates=['FECHA_VISADO'])

        # Crear archivo con los datos filtrados por el mes seleccionado
        archivo.seek(0)
        archivo_bytes = BytesIO(archivo.read())
        libro = load_workbook(archivo_bytes)

        # El mes seleccionado es el nombre del mes. Para convertirlo en el número del mes, usamos calendar.month_name
        mes_num = list(calendar.month_name[1:]).index(mes_seleccionado) + 1  # Obtiene el índice del mes (1-12)
        
        # Llamar a la función para generar las hojas con el mes seleccionado
        crear_hoja_mes_seleccionado(libro, f"DTO_{mes_seleccionado}", df_dto, mes_num)
        crear_hoja_mes_seleccionado(libro, f"PCL_{mes_seleccionado}", df_pcl, mes_num)

        # Llamar a la función para generar las tablas de DTO y PCL
        generar_tablas_dto_y_pcl(libro, df_dto, df_pcl)

        # Llamar a la función para crear la hoja de comparativa de año
        crear_comparativa_ano(libro, df_dto)  # También puedes pasar df_pcl si es necesario

        output = BytesIO()
        libro.save(output)
        output.seek(0)
        descargar_archivo(output, nombre="informe_dto_pcl_mes.xlsx")
        st.success("✅ Archivo generado con éxito.")
    elif archivo and tipo == "csv":
        st.warning("Actualmente el procesamiento está disponible solo para archivos .xlsx con hojas DTO y PCL.")